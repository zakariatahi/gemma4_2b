import os
import sys
import json
import uuid
import datetime
import requests
import openpyxl
import atexit
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from duckduckgo_search import DDGS

from langchain_ollama import ChatOllama
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.tools import tool
from langchain_classic.agents import AgentExecutor, create_tool_calling_agent
from langchain_core.messages import HumanMessage, AIMessage
from langchain_core.callbacks.base import BaseCallbackHandler


# ─────────────────────────────────────────
# 1. CLASSES: PROFILE & STORAGE  (in-memory caches, lazy flush)
# ─────────────────────────────────────────

class UserProfile:
    """
    OPTIMIZATION: Keeps a live in-memory dict.
    - Reads disk ONCE at startup, then works purely from memory.
    - Writes to disk only when data actually changes (write-through).
    - No per-message JSON load.
    """
    def __init__(self, file_path: str = "user_profile.json"):
        self.file_path = file_path
        self._cache: dict = {}
        self._load_once()

    def _load_once(self):
        if os.path.exists(self.file_path):
            with open(self.file_path, "r", encoding="utf-8") as f:
                self._cache = json.load(f)

    def _flush(self):
        with open(self.file_path, "w", encoding="utf-8") as f:
            json.dump(self._cache, f, indent=2, ensure_ascii=False)

    def update(self, key: str, value: str):
        self._cache[key] = value
        self._flush()
        print(f"\n🔍 DEBUG: Profile saved → {key}: {value}")

    def delete(self, key: str):
        if key in self._cache:
            del self._cache[key]
            self._flush()

    def get_all(self) -> dict:
        return self._cache          # no disk I/O

    def to_prompt_string(self) -> str:
        if not self._cache:
            return "No information about the user yet."
        return "\n".join(f"- {k}: {v}" for k, v in self._cache.items())


class ConversationStorage:
    """
    OPTIMIZATION: Buffers messages in memory; flushes to disk every
    FLUSH_EVERY messages (default 5) or on clean exit.
    - Eliminates per-message JSON load+write.
    - Still persists correctly even if you Ctrl-C (atexit flush).
    """
    FLUSH_EVERY = 5

    def __init__(self, file_path: str = "coach_conversations.json"):
        self.file_path = file_path
        self.session_id = str(uuid.uuid4())[:8]
        self._session_buf: dict = {
            "session_id": self.session_id,
            "date": datetime.now().strftime("%Y-%m-%d"),
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "messages": []
        }
        self._unflushed = 0
        atexit.register(self._flush)

    def _flush(self):
        if not self._session_buf["messages"]:
            return
        if os.path.exists(self.file_path):
            with open(self.file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        else:
            data = {"sessions": []}

        existing = next(
            (s for s in data["sessions"] if s["session_id"] == self.session_id), None
        )
        if existing:
            existing.update(self._session_buf)
        else:
            data["sessions"].append(self._session_buf)

        self._session_buf["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._session_buf["total_messages"] = len(self._session_buf["messages"])

        with open(self.file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        self._unflushed = 0

    def save_message(self, role: str, message: str):
        self._session_buf["messages"].append({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "role": role,
            "message": message
        })
        self._unflushed += 1
        if self._unflushed >= self.FLUSH_EVERY:
            self._flush()

    def print_history(self):
        msgs = self._session_buf["messages"]
        if not msgs:
            print("No messages yet.")
            return
        print(f"\n📅 Session: {self.session_id} — {self._session_buf['date']}")
        print("=" * 55)
        for msg in msgs:
            role = "You" if msg["role"] == "human" else "COACH"
            print(f"[{msg['timestamp']}] {role}: {msg['message']}\n")


# ─────────────────────────────────────────
# 2. HANDLERS & LOADERS
# ─────────────────────────────────────────

class StreamingHandler(BaseCallbackHandler):
    def on_llm_new_token(self, token: str, **kwargs):
        print(token, end="", flush=True)

    def on_tool_start(self, serialized, input_str, **kwargs):
        print(f"\n🔧 Using tool : {serialized.get('name', 'unknown')}")
        print(f"   ↳ Input     : {input_str}\n")

    def on_tool_end(self, output, **kwargs):
        preview = str(output)[:400]
        print(f"📥 Tool result : {preview}{'...' if len(str(output)) > 400 else ''}")
        print("\n🧠 Thinking ", end="", flush=True)


def load_document(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".txt":
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    elif ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        return "\n".join(page.extract_text() for page in reader.pages)
    else:
        raise ValueError("Only .pdf and .txt files are supported.")


# ─────────────────────────────────────────
# 3. TOOLS  (same logic; closures created once)
# ─────────────────────────────────────────

def make_profile_tools(profile: UserProfile):
    @tool
    def update_user_profile(key: str, value: str) -> str:
        """Updates user profile with info like name, job, hobbies, or goals."""
        profile.update(key, value)
        return f"✅ Got it! I'll remember that {key}: {value}"

    @tool
    def forget_user_info(key: str) -> str:
        """Removes a piece of information from the user profile."""
        profile.delete(key)
        return f"✅ Got it! I've forgotten: {key}"

    @tool
    def show_user_profile() -> str:
        """Shows everything COACH knows about the user."""
        data = profile.get_all()
        if not data:
            return "I don't know anything about you yet!"
        lines = "\n".join(f"  • {k}: {v}" for k, v in data.items())
        return f"Here's what I know about you:\n{lines}"

    return [update_user_profile, forget_user_info, show_user_profile]


@tool
def get_current_time() -> str:
    """Returns the current date and time."""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

@tool
def search_web(query: str) -> str:
    """Searches the web for productivity tips or health advice."""
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=3))
            return (
                "\n\n".join(f"{i+1}. {r['title']}\n{r['body']}" for i, r in enumerate(results))
                if results else "No results found."
            )
    except Exception as e:
        return f"Search error: {e}"

@tool
def calculator(expression: str) -> str:
    """Evaluates a math expression."""
    try:
        return str(eval(expression, {"__builtins__": {}}))
    except Exception as e:
        return f"Error: {e}"

@tool
def get_weather(city: str) -> str:
    """Gets current weather for a city."""
    try:
        response = requests.get(f"https://wttr.in/{city}?format=3", timeout=5)
        return response.text
    except Exception as e:
        return f"Error: {e}"

@tool
def save_note(note: str) -> str:
    """Saves an important note or recommendation to a file."""
    with open("coach_notes.txt", "a", encoding="utf-8") as f:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        f.write(f"[{timestamp}] {note}\n")
    return "Note saved successfully!"

@tool
def rate_my_day(
    productive_hours: float, wasted_hours: float,
    sleep_hours: float, exercise_minutes: float
) -> str:
    """Calculates a productivity score out of 100."""
    score = 0
    score += min(productive_hours * 10, 40)
    score += max(0, 20 - wasted_hours * 5)
    score += 20 if 7 <= sleep_hours <= 9 else 10 if sleep_hours >= 6 else 0
    score += min(exercise_minutes / 3, 20)
    grade = (
        "🟢 Excellent" if score >= 80
        else "🟡 Good" if score >= 60
        else "🔴 Needs Work"
    )
    return f"Your day score: {score:.0f}/100 — {grade}"

@tool
def get_nutrition_advice(meal_description: str) -> str:
    """Searches for nutrition advice based on what the user ate."""
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(f"nutrition advice {meal_description}", max_results=2))
            return "\n".join(r["body"] for r in results)
    except Exception as e:
        return f"Error: {e}"

@tool
def analyze_sleep(bedtime: str, wake_time: str) -> str:
    """Analyzes sleep quality. Format: HH:MM (24h)."""
    fmt = "%H:%M"
    bed = datetime.strptime(bedtime, fmt)
    wake = datetime.strptime(wake_time, fmt)
    if wake < bed:
        wake += timedelta(days=1)
    duration = (wake - bed).seconds / 3600
    quality = (
        "✅ Optimal" if 7 <= duration <= 9
        else "⚠️ Slightly low" if duration >= 6
        else "🔴 Insufficient"
    )
    return f"Sleep duration: {duration:.1f} hours — {quality}"

@tool
def recommend_exercise(available_minutes: int, fitness_goal: str) -> str:
    """Recommends a workout based on available time and goal."""
    try:
        with DDGS() as ddgs:
            query = f"{available_minutes} minute workout for {fitness_goal}"
            results = list(ddgs.text(query, max_results=2))
            return "\n".join(r["body"] for r in results)
    except Exception as e:
        return f"Error: {e}"

@tool
def create_tomorrow_plan(tasks: str) -> str:
    """Saves a prioritized plan for tomorrow to a file."""
    task_list = [t.strip() for t in tasks.split(",")]
    with open("tomorrow_plan.txt", "w", encoding="utf-8") as f:
        f.write("📅 Tomorrow's Plan\n" + "=" * 30 + "\n")
        for i, task in enumerate(task_list, 1):
            f.write(f"{i}. {task}\n")
    return f"Plan saved with {len(task_list)} tasks."

@tool
def save_to_excel(
    date: str = "unknown", wake_time: str = "unknown", sleep_time: str = "unknown",
    productive_hours: float = 0.0, wasted_hours: float = 0.0,
    exercise_minutes: int = 0, meals_count: int = 0,
    productivity_stars: int = 3, mood: str = "unknown", notes: str = ""
) -> str:
    """Saves the day's summary to an Excel file."""
    file_path = "daily_log.xlsx"
    wb = openpyxl.load_workbook(file_path) if os.path.exists(file_path) else openpyxl.Workbook()
    ws = wb.active
    if not os.path.exists(file_path):
        ws.title = "Daily Log"
        headers = [
            "📅 Date", "⏰ Wake Up", "🌙 Sleep", "✅ Productive Hrs",
            "❌ Wasted Hrs", "🏃 Exercise (min)", "🍽️ Meals",
            "⭐ Productivity", "😊 Mood", "📝 Notes"
        ]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill("solid", fgColor="2E86AB")
            cell.font = Font(bold=True, color="FFFFFF")

    stars = "⭐" * productivity_stars + "☆" * (5 - productivity_stars)
    row_data = [
        date, wake_time, sleep_time, productive_hours, wasted_hours,
        exercise_minutes, meals_count, stars, mood, notes
    ]
    existing_row = next(
        (row[0].row for row in ws.iter_rows(min_row=2) if str(row[0].value) == date), None
    )
    if existing_row:
        for col, val in enumerate(row_data, 1):
            ws.cell(row=existing_row, column=col, value=val)
    else:
        ws.append(row_data)

    wb.save(file_path)
    return f"✅ Logged to Excel for {date}."


tools_list = [
    get_current_time, search_web, calculator, save_note,
    rate_my_day, analyze_sleep, recommend_exercise,
    create_tomorrow_plan, save_to_excel, get_weather, get_nutrition_advice
]


# ─────────────────────────────────────────
# 4. PROMPT  (built ONCE; profile string injected per-turn via {user_profile})
# ─────────────────────────────────────────

def build_prompt(log_content: str) -> ChatPromptTemplate:
    """
    OPTIMIZATION: log_content is static for the session, so the template
    is built once.  The live profile string is passed as a regular input
    key each turn — no template rebuild needed.
    """
    return ChatPromptTemplate.from_messages([
        ("system", f"""You are COACH, a personal productivity AI.

USER PROFILE:
{{user_profile}}

ACTIVITY LOG:
{log_content}

CRITICAL INSTRUCTION:
If the user mentions a name, job, goal, city, or any new fact about themselves,
you MUST call 'update_user_profile' BEFORE answering.
Do not just say you will remember it—actually use the tool.

STEP-BY-STEP PROCESS:
1. Scan user input for personal info.
2. If found, call update_user_profile.
3. Then, provide your coaching response.

Intent detection:
1. FULL ANALYSIS  → run sleep, rate, plan, excel tools.
2. SPECIFIC QUESTION → direct answer from log.
3. SAVE TO EXCEL → call save_to_excel immediately.
"""),
        ("placeholder", "{chat_history}"),
        ("human", "{input}"),
        ("placeholder", "{agent_scratchpad}"),
    ])


# ─────────────────────────────────────────
# 5. MAIN  — executor built ONCE, reused every turn
# ─────────────────────────────────────────

def main():
    profile = UserProfile()

    file_path = sys.argv[1] if len(sys.argv) > 1 else input("📂 Path to log: ").strip()
    log_content = load_document(file_path)

    storage = ConversationStorage()
    all_tools = tools_list + make_profile_tools(profile)
    handler = StreamingHandler()

    # ── Model recommendation ─────────────────────────────────────────────
    # CPU-only?  Prefer a 1-2 B model.  Ranked by speed on CPU (fastest first):
    #   qwen2.5:1.5b   ← fastest, surprisingly capable for short coaching tasks
    #   gemma2:2b
    #   phi3:mini
    #   mistral:7b-q4  ← if you have ≥16 GB RAM and don't mind waiting
    #
    # Change the model name below to whatever you have pulled in Ollama.
    # ─────────────────────────────────────────────────────────────────────
    MODEL = "gemma4:e2b"

    llm = ChatOllama(
        model=MODEL,
        streaming=True,
        callbacks=[handler],
        # These reduce CPU memory pressure and improve throughput:
        num_ctx=2048,       # smaller context window = less RAM, faster prefill
        num_thread=max(1, os.cpu_count() - 1),  # leave 1 core for the OS
    )

    # OPTIMIZATION: build prompt and executor ONCE, not per turn
    prompt = build_prompt(log_content)
    agent = create_tool_calling_agent(llm, all_tools, prompt)
    executor = AgentExecutor(
        agent=agent,
        tools=all_tools,
        callbacks=[handler],
        verbose=False,
        max_iterations=6,   # guard against runaway tool loops
    )

    chat_history: list = []

    print("\n" + "=" * 55 + f"\n COACH ONLINE  [{MODEL}]\n" + "=" * 55)

    while True:
        try:
            user_input = input("\nYou: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n\n👋 Goodbye!")
            break

        if not user_input:
            continue
        if user_input.lower() == "quit":
            break
        if user_input.lower() == "history":
            storage.print_history()
            continue

        try:
            storage.save_message("human", user_input)

            response = executor.invoke({
                "input": user_input,
                "chat_history": chat_history,
                "user_profile": profile.to_prompt_string(),   # ← injected here, no rebuild
            })
            answer = response["output"]

            print(f"\nCOACH: {answer}")
            storage.save_message("coach", answer)

            # Keep only the last 10 turns in memory to limit context size
            chat_history.extend([
                HumanMessage(content=user_input),
                AIMessage(content=answer),
            ])
            if len(chat_history) > 20:
                chat_history = chat_history[-20:]

        except Exception as e:
            print(f"\nCOACH Error: {e}")


if __name__ == "__main__":
    main()